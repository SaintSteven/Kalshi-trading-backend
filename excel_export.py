from __future__ import annotations

from io import BytesIO
from datetime import datetime, timezone
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from models import ExportCardRequest


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
EDGE_FILL = PatternFill("solid", fgColor="E2F0D9")
WATCH_FILL = PatternFill("solid", fgColor="FFF2CC")
PASS_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="B7C9DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)



def _game_details(ticker: str) -> tuple[str, str]:
    """Return a readable matchup and scheduled ET time from a Kalshi MLB ticker."""
    parts = (ticker or "").split("-")
    event = parts[1] if len(parts) > 1 else ""
    match = re.fullmatch(r"\d{2}[A-Z]{3}\d{2}(\d{4})([A-Z]{6})", event, re.IGNORECASE)
    if not match:
        return "Game unavailable", "Time unavailable"
    military_time, matchup = match.groups()
    matchup = matchup.upper()
    away, home = matchup[:3], matchup[3:]
    hour = int(military_time[:2])
    minute = military_time[2:]
    suffix = "PM" if hour >= 12 else "AM"
    hour_12 = ((hour + 11) % 12) + 1
    return f"{away} @ {home}", f"{hour_12}:{minute} {suffix} ET"


def _confidence(rec) -> float | None:
    value = (rec.confidence or {}).get("overall")
    if value is None:
        return None
    value = float(value)
    return value / 100 if value > 1 else value


def _implied_probability(rec) -> float | None:
    if rec.market_price_cents is None or rec.side == "NONE":
        return None
    return rec.market_price_cents / 100


def _buy_range(rec) -> str:
    if rec.market_price_cents is None or rec.side == "NONE":
        return "—"
    return f"Buy {rec.side} ≤ {rec.market_price_cents}¢"


def _qc_status(rec) -> str:
    if rec.decision == "INSUFFICIENT DATA":
        return "FAIL"
    if rec.warnings:
        return "REVIEW"
    if rec.decision == "MODEL EDGE":
        return "PRELIMINARY PASS"
    return "NO ACTION"


def build_card_workbook(payload: ExportCardRequest) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = payload.card_date or "Daily Card"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Kalshi MLB Paper Trading Card"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:T1")

    metadata = [
        ("Card Date", payload.card_date or "Current slate"),
        ("Generated At", payload.generated_at or datetime.now(timezone.utc).isoformat()),
        ("Model Version", payload.model_version),
        ("Starting Bankroll", payload.bankroll),
        ("Already Committed", payload.already_committed_today),
        ("Selected Slate", payload.selected_slate or ""),
    ]
    for row, (label, value) in enumerate(metadata, start=3):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = SUB_FILL
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2, value=value).border = BORDER

    header_row = 10
    headers = [
        "Player", "Game", "Start Time", "Ticker", "Side", "Ladder", "Projection Ks",
        "Market Price (¢)", "Implied %", "Fair %", "Raw Edge (pts)",
        "Adjusted Edge (pts)", "Confidence", "Stake ($)", "Buy Range",
        "QC Status", "Actual Ks", "Result", "Paper P/L ($)", "Notes"
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    recommendations = payload.recommendations
    first_data_row = header_row + 1
    for row_idx, rec in enumerate(recommendations, start=first_data_row):
        notes = " | ".join([*(rec.reasons or []), *(rec.warnings or [])])
        fallback_game, fallback_start_time = _game_details(rec.ticker)
        game = rec.matchup or fallback_game
        start_time = rec.game_start_display or fallback_start_time
        values = [
            rec.player, game, start_time, rec.ticker, rec.side, rec.threshold,
            rec.projected_strikeouts, rec.market_price_cents, _implied_probability(rec),
            rec.fair_probability, rec.raw_edge_points, rec.adjusted_edge_points,
            _confidence(rec), rec.suggested_stake, _buy_range(rec), _qc_status(rec),
            None, None, None, notes,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 20))

        fill = EDGE_FILL if rec.decision == "MODEL EDGE" else WATCH_FILL if rec.decision == "WATCH" else PASS_FILL
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

        for col_idx in (9, 10, 13):
            ws.cell(row=row_idx, column=col_idx).number_format = "0.0%"
        ws.cell(row=row_idx, column=14).number_format = '$0.00'
        ws.cell(row=row_idx, column=19).number_format = '$0.00'

    # Excel for iOS can flag a workbook as damaged when an OOXML table is
    # created over a placeholder-only row.  Only create a real Excel table
    # when the card actually contains recommendations.  Empty cards remain a
    # valid, formatted journal sheet without a table object.
    if recommendations:
        last_data_row = first_data_row + len(recommendations) - 1
        table = Table(displayName="DailyCard", ref=f"A{header_row}:T{last_data_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=False, showColumnStripes=False,
        )
        ws.add_table(table)
        ws.auto_filter.ref = f"A{header_row}:T{last_data_row}"
    else:
        last_data_row = first_data_row
        ws.merge_cells(start_row=first_data_row, start_column=1,
                       end_row=first_data_row, end_column=len(headers))
        empty_cell = ws.cell(row=first_data_row, column=1, value="No recommendations returned for this slate.")
        empty_cell.fill = WATCH_FILL
        empty_cell.font = Font(italic=True)
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")
        empty_cell.border = BORDER
        ws.row_dimensions[first_data_row].height = 24

    summary_row = last_data_row + 3
    ws.cell(summary_row, 1, "Daily Summary").font = Font(size=13, bold=True)
    summary = [
        ("Recommendations", len(recommendations)),
        ("Model Edge Bets", sum(1 for r in recommendations if r.decision == "MODEL EDGE" and r.suggested_stake > 0)),
        ("Total Planned Stake", sum(r.suggested_stake for r in recommendations)),
        ("Wins", f'=COUNTIF(R{first_data_row}:R{last_data_row},"W")'),
        ("Losses", f'=COUNTIF(R{first_data_row}:R{last_data_row},"L")'),
        ("Paper P/L", f'=SUM(S{first_data_row}:S{last_data_row})'),
        ("Ending Bankroll", f'=B6+B{summary_row + 6}'),
    ]
    for idx, (label, value) in enumerate(summary, start=summary_row + 1):
        ws.cell(idx, 1, label).font = Font(bold=True)
        ws.cell(idx, 1).fill = SUB_FILL
        ws.cell(idx, 1).border = BORDER
        ws.cell(idx, 2, value).border = BORDER
    ws.cell(summary_row + 3, 2).number_format = '$0.00'
    ws.cell(summary_row + 6, 2).number_format = '$0.00'
    ws.cell(summary_row + 7, 2).number_format = '$0.00'

    widths = [22, 16, 16, 26, 9, 10, 13, 16, 12, 12, 15, 19, 13, 12, 18, 18, 11, 10, 15, 60]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = f"A{first_data_row}"
    ws.print_title_rows = f"1:{header_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    instructions = wb.create_sheet("Instructions")
    instructions["A1"] = "Paper Testing Workflow"
    instructions["A1"].font = Font(size=15, bold=True)
    steps = [
        "1. Export immediately after building the card and before first pitch.",
        "2. Copy this daily sheet into the master workbook.",
        "3. Do not alter recommendation fields after export.",
        "4. After games, enter Actual Ks, Result (W/L), Paper P/L, and optional notes.",
        "5. QC Status is preliminary until lineup, workload, weather, and other late news are reviewed.",
    ]
    for row_idx, step in enumerate(steps, start=3):
        instructions.cell(row_idx, 1, step)
    instructions.column_dimensions["A"].width = 110

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
